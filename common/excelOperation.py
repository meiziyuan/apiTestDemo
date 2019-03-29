import openpyxl
import os, sys


##os.path.split(os.path.realpath(__file__))[0] 这个用法才是获取当前脚本的目录，os.getcwd()和sys.path[0]都获取的是python脚本运行的路径。
file = os.path.dirname(os.path.split(os.path.realpath(__file__))[0])+"\\testfiles\\Api1.xlsx"

class excelAPI():

    def __init__(self, sheetname):
        self.sheetname = sheetname
        self.workbook = openpyxl.load_workbook(file)

    def getCaseData(self, caseid):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[self.sheetname]

        info = {}
        for i in range(6, 100):
            if sheet.cell(row=i, column=2).value == caseid:
                info["apiName"] = sheet.cell(row=i, column=1).value
                info["caseName"] = sheet.cell(row=i, column=3).value
                info["url"] = sheet.cell(row=i, column=4).value
                info["method"] = sheet.cell(row=i, column=5).value
                info["params"] = eval(sheet.cell(row=i, column=6).value)
                info["params"]["model"] = sheet.cell(row=i, column=7).value
                info["params"]["action"] = sheet.cell(row=i, column=8).value
                break
        return info

    def closeXsl(self):
        self.workbook.close()


if __name__ == "__main__":
    dd = excelAPI('GetRoomListApi')
    print(dd.getCaseData('001'))
    aa = dd.getCaseData('007')['params']
    dd.closeXsl()
    print(aa, type(aa))